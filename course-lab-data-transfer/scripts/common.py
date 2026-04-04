#!/usr/bin/env python3
"""Shared helpers for the course-lab-data-transfer skill."""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path


IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}
TEXT_SUFFIXES = {".txt", ".text"}
MARKDOWN_SUFFIXES = {".md", ".markdown"}
DELIMITED_SUFFIXES = {".csv", ".tsv"}
SPREADSHEET_SUFFIXES = {".xls", ".xlsx"}


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def detect_source_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return "pdf"
    if suffix in IMAGE_SUFFIXES:
        return "image"
    if suffix == ".csv":
        return "csv"
    if suffix == ".tsv":
        return "tsv"
    if suffix in TEXT_SUFFIXES:
        return "text"
    if suffix in MARKDOWN_SUFFIXES:
        return "markdown"
    if suffix in SPREADSHEET_SUFFIXES:
        return "spreadsheet"
    return "unknown"


def markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    header = normalized[0]
    separator = ["---"] * width
    table_rows = [header, separator, *normalized[1:]]
    return "\n".join("| " + " | ".join(cell or "" for cell in row) + " |" for row in table_rows)


def read_delimited_preview(path: Path, *, delimiter: str, max_rows: int) -> str:
    rows: list[list[str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for index, row in enumerate(reader):
            if index >= max_rows:
                break
            rows.append([str(cell).strip() for cell in row])
    table = markdown_table(rows)
    if not table:
        return f"# Direct Preview\n\nSource: `{path}`\n\nNo non-empty rows were found.\n"
    return "\n".join(
        [
            "# Direct Preview",
            "",
            f"Source: `{path}`",
            "",
            table,
            "",
        ]
    )


def read_text_preview(path: Path) -> str:
    return "\n".join(
        [
            "# Direct Preview",
            "",
            f"Source: `{path}`",
            "",
            path.read_text(encoding="utf-8"),
            "",
        ]
    )


def read_workbook_preview(path: Path, *, max_rows: int) -> tuple[str | None, str | None]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, "openpyxl is required to preview spreadsheet sources."

    workbook = load_workbook(path, read_only=True, data_only=True)
    lines = ["# Spreadsheet Preview", "", f"Source: `{path}`", ""]
    for sheet in workbook.worksheets:
        lines.append(f"## Sheet: {sheet.title}")
        rows: list[list[str]] = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= max_rows:
                break
            rows.append(["" if value is None else str(value) for value in row])
        table = markdown_table(rows)
        if table:
            lines.append(table)
        else:
            lines.append("No non-empty preview rows found.")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n", None


def pretty_experiment_name(value: str) -> str:
    stripped = value.strip()
    if not stripped:
        return "Experiment"
    if re.search(r"[A-Za-z]", stripped):
        return stripped.replace("_", " ").replace("-", " ").title()
    return stripped
