#!/usr/bin/env python3
"""Compute type A/B/C and expanded uncertainties from normalized measurement tables."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import statistics
import unicodedata
from pathlib import Path

from common import write_json


NUMBER_RE = re.compile(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?")
GREEK_NAME_MAP = {
    "α": "alpha",
    "β": "beta",
    "γ": "gamma",
    "δ": "delta",
    "Δ": "Delta",
    "ε": "epsilon",
    "θ": "theta",
    "κ": "kappa",
    "λ": "lambda",
    "μ": "mu",
    "ν": "nu",
    "π": "pi",
    "ρ": "rho",
    "σ": "sigma",
    "φ": "phi",
    "ω": "omega",
}
DIGIT_WORDS = {
    "0": "zero",
    "1": "one",
    "2": "two",
    "3": "three",
    "4": "four",
    "5": "five",
    "6": "six",
    "7": "seven",
    "8": "eight",
    "9": "nine",
}


def parse_number(raw: str) -> float | None:
    match = NUMBER_RE.search(raw.replace(",", ""))
    if not match:
        return None
    return float(match.group(0))


def parse_quantity_label(raw_label: str) -> dict[str, str | None]:
    label = raw_label.strip()
    if "/" in label:
        left, right = label.rsplit("/", 1)
        left = left.strip()
        right = right.strip()
        if left and right and len(right) <= 32:
            return {
                "raw_label": label,
                "quantity_label": left,
                "unit": right,
            }
    return {
        "raw_label": label,
        "quantity_label": label,
        "unit": None,
    }


def canonical_symbol_key(symbol: str) -> str:
    text = unicodedata.normalize("NFKC", symbol).strip()
    text = text.replace("(", "").replace(")", "")
    text = text.replace("'", "_prime").replace("′", "_prime").replace("″", "_double_prime")
    for greek, name in GREEK_NAME_MAP.items():
        text = text.replace(greek, name)

    leading_digits = ""
    while text and text[0].isdigit():
        leading_digits += text[0]
        text = text[1:]
    if leading_digits:
        digit_prefix = "_".join(DIGIT_WORDS[digit] for digit in leading_digits)
        text = f"{digit_prefix}_{text}" if text else digit_prefix

    text = re.sub(r"[^0-9A-Za-z_]+", "_", text)
    text = re.sub(r"_{2,}", "_", text).strip("_")
    if not text:
        return "quantity"
    if text[0].isdigit():
        text = f"quantity_{text}"
    return text


def read_csv_like(path: Path, delimiter: str | None) -> list[dict[str, str]]:
    sample = path.read_text(encoding="utf-8", errors="replace")
    used_delimiter = delimiter
    if used_delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(sample[:4096], delimiters=",\t;|")
            used_delimiter = dialect.delimiter
        except csv.Error:
            used_delimiter = "\t" if path.suffix.lower() in {".tsv", ".txt"} else ","
    with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=used_delimiter))


def read_markdown_table(path: Path) -> list[dict[str, str]]:
    lines = [line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    rows = [line for line in lines if line.startswith("|") and line.endswith("|")]
    if len(rows) < 2:
        raise SystemExit(f"No markdown table found in {path}")
    headers = [item.strip() for item in rows[0].strip("|").split("|")]
    data_rows: list[dict[str, str]] = []
    for raw in rows[2:]:
        values = [item.strip() for item in raw.strip("|").split("|")]
        if len(values) != len(headers):
            continue
        data_rows.append(dict(zip(headers, values)))
    return data_rows


def read_json_rows(path: Path) -> list[dict[str, str]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list) or not all(isinstance(item, dict) for item in payload):
        raise SystemExit(f"JSON input must be a list of objects: {path}")
    return [{str(key): str(value) for key, value in item.items()} for item in payload]


def read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(f"openpyxl is required for {path.suffix} inputs: {exc}") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() for value in rows[0]]
    payload: list[dict[str, str]] = []
    for row in rows[1:]:
        payload.append({headers[idx]: "" if value is None else str(value) for idx, value in enumerate(row)})
    return payload


def load_rows(path: Path, delimiter: str | None) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        return read_csv_like(path, delimiter)
    if suffix == ".md":
        return read_markdown_table(path)
    if suffix == ".json":
        return read_json_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    raise SystemExit(f"Unsupported input format for uncertainty computation: {path}")


def resolution_map(raw_items: list[str]) -> dict[str, float]:
    payload: dict[str, float] = {}
    for item in raw_items:
        name, _, value = item.partition("=")
        if not name or not value:
            raise SystemExit(f"Invalid --resolution value: {item}. Expected column=value")
        payload[name.strip()] = float(value.strip())
    return payload


def auto_numeric_columns(rows: list[dict[str, str]]) -> list[str]:
    if not rows:
        return []
    picked: list[str] = []
    for column in rows[0].keys():
        values = [parse_number(str(row.get(column, ""))) for row in rows]
        numbers = [value for value in values if value is not None]
        if len(numbers) >= 2:
            picked.append(column)
    return picked


def summarize_column(values: list[float], resolution: float | None, coverage_k: float) -> dict[str, float | int]:
    mean = statistics.fmean(values)
    if len(values) > 1:
        sample_std = statistics.stdev(values)
        type_a = sample_std / math.sqrt(len(values))
    else:
        sample_std = 0.0
        type_a = 0.0
    type_b = (resolution / math.sqrt(3)) if resolution is not None else 0.0
    type_c = math.sqrt(type_a ** 2 + type_b ** 2)
    expanded = coverage_k * type_c
    return {
        "n": len(values),
        "mean": mean,
        "sample_stddev": sample_std,
        "type_a": type_a,
        "resolution": resolution,
        "type_b": type_b,
        "type_c": type_c,
        "expanded_uncertainty": expanded,
    }


def markdown_report(input_path: Path, columns: list[tuple[str, dict[str, float | int]]], coverage_k: float) -> str:
    lines = [
        "# Processed Uncertainty Summary",
        "",
        f"- Source table: `{input_path}`",
        f"- Coverage factor `k`: {coverage_k:g}",
        "",
        "| Quantity | Unit | n | Mean | Sample Std. Dev. | Type A | Resolution | Type B | Type C | Expanded |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for name, stats in columns:
        parsed = parse_quantity_label(name)
        display_resolution = "TBD" if stats["resolution"] is None else f"{stats['resolution']:.6g}"
        lines.append(
            f"| {parsed['quantity_label']} | {parsed['unit'] or '-'} | {stats['n']} | {stats['mean']:.6g} | {stats['sample_stddev']:.6g} | "
            f"{stats['type_a']:.6g} | {display_resolution} | {stats['type_b']:.6g} | "
            f"{stats['type_c']:.6g} | {stats['expanded_uncertainty']:.6g} |"
        )
    lines.append("")
    lines.append("Type B uses `resolution / sqrt(3)` when a resolution is provided.")
    return "\n".join(lines) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--quantity", action="append", default=[])
    parser.add_argument("--resolution", action="append", default=[], help="Repeatable column=value entry.")
    parser.add_argument("--coverage-k", type=float, default=2.0)
    parser.add_argument("--delimiter", help="Override delimiter for CSV/TSV/TXT inputs.")
    parser.add_argument("--output-markdown", required=True)
    parser.add_argument("--output-json")
    args = parser.parse_args()

    input_path = Path(args.input)
    rows = load_rows(input_path, args.delimiter)
    if not rows:
        raise SystemExit(f"No rows found in {input_path}")

    picked_columns = args.quantity or auto_numeric_columns(rows)
    if not picked_columns:
        raise SystemExit("No numeric columns found. Pass --quantity for the measurement columns to summarize.")

    resolutions = resolution_map(args.resolution)
    summaries: list[tuple[str, dict[str, float | int]]] = []
    json_columns: dict[str, dict[str, float | int]] = {}
    for column in picked_columns:
        values = [parse_number(str(row.get(column, ""))) for row in rows]
        numeric = [value for value in values if value is not None]
        if len(numeric) < 1:
            raise SystemExit(f"Column '{column}' does not contain numeric values.")
        summary = summarize_column(numeric, resolutions.get(column), args.coverage_k)
        summary.update(parse_quantity_label(column))
        summary["canonical_key"] = canonical_symbol_key(str(summary["quantity_label"]))
        summaries.append((column, summary))
        json_columns[column] = summary

    markdown = markdown_report(input_path, summaries, args.coverage_k)
    Path(args.output_markdown).write_text(markdown, encoding="utf-8")
    if args.output_json:
        write_json(
            Path(args.output_json),
            {
                "input": str(input_path),
                "coverage_k": args.coverage_k,
                "columns": json_columns,
            },
        )

    print(json.dumps({"output_markdown": args.output_markdown, "output_json": args.output_json}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
